import json  # Thư viện xử lý JSON
from openpyxl import load_workbook  # Thư viện đọc file Excel .xlsx

def load_test_cases_from_excel(file_path):
    # Mở file Excel test case
    wb = load_workbook(file_path)

    # Lấy sheet đầu tiên (sheet mặc định)
    sheet = wb.active

    # Lấy tiêu đề các cột từ hàng đầu tiên
    headers = [cell.value for cell in sheet[1]]

    # Khởi tạo danh sách để chứa các test case
    test_cases = []

    # Duyệt qua từng dòng dữ liệu, bắt đầu từ hàng 2 (bỏ qua tiêu đề)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Ghép tiêu đề với giá trị từng cột thành dictionary
        row_data = dict(zip(headers, row))

        # Nếu có dữ liệu trong request_body, parse từ chuỗi JSON sang dict
        row_data['request_body'] = json.loads(row_data['request_body']) if row_data['request_body'] else {}

        # Nếu có dữ liệu trong expected_response, parse từ chuỗi JSON sang dict
        row_data['expected_response'] = json.loads(row_data['expected_response']) if row_data['expected_response'] else {}

        # Thêm test case vào danh sách
        test_cases.append(row_data)
    
    # Trả về danh sách các test case đã đọc được
    return test_cases
