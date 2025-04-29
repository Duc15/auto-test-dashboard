import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from datetime import datetime

def save_test_results(test_cases, results, file_name='test_result.xlsx'):
    records = []
    
    # Lấy thời gian hiện tại (theo định dạng YYYY-MM-DD HH:MM:SS)
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Tạo danh sách các bản ghi kiểm thử để lưu vào Excel
    for case, result in zip(test_cases, results):
        record = {
            'api_name': case['api_name'],                    # Tên API
            'method': case['method'],                        # Phương thức HTTP (GET, POST, ...)
            'endpoint': case['endpoint'],                    # Đường dẫn API
            'request_body': case['request_body'],            # Nội dung request gửi lên
            'expected_status': case['expected_status'],      # Mã trạng thái mong đợi
            'expected_response': case['expected_response'],  # Kết quả mong đợi
            'actual_response': result['actual_response'],    # Kết quả thực tế nhận được
            'result': result['result'],                      # Kết quả kiểm thử (Pass/Fail)
            'message': result['message']                     # Thông điệp hoặc lý do nếu Fail
        }
        records.append(record)  # Thêm bản ghi vào danh sách

    # Chuyển danh sách bản ghi thành DataFrame để xử lý với pandas
    df = pd.DataFrame(records)

    # Dùng ExcelWriter để ghi dữ liệu vào file Excel
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        # Ghi DataFrame vào sheet "Test Results", không ghi cột index
        df.to_excel(writer, sheet_name='Test Results', index=False)

        # Tạo sheet "Summary" chứa tổng quan kết quả kiểm thử
        summary_data = {
            'Total Tests': [len(test_cases)],                       # Tổng số test case
            'Pass': [sum(1 for result in results if result['result'] == 'Pass')],  # Số lượng Pass
            'Fail': [sum(1 for result in results if result['result'] == 'Fail')],  # Số lượng Fail
            'Test Date': [current_date]                             # Ngày và giờ kiểm thử
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    # Mở file Excel vừa tạo để thêm định dạng (border)
    wb = load_workbook(file_name)
    ws = wb['Test Results']       # Sheet kết quả chi tiết
    summary_ws = wb['Summary']    # Sheet tổng quan

    # Tạo khung viền mỏng cho ô (thin border)
    thin_border = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin'),
        left=Side(style='thin'),
        right=Side(style='thin')
    )

    # Áp dụng border cho tất cả ô dữ liệu trong sheet "Test Results"
    for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.border = thin_border

    # Áp dụng border cho tất cả ô dữ liệu trong sheet "Summary"
    for row in summary_ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=len(summary_data)):
        for cell in row:
            cell.border = thin_border

    # Lưu lại workbook sau khi đã thêm border
    wb.save(file_name)
    
    # In ra thông báo đã lưu file
    print(f"📁 Đã lưu kết quả test vào: {file_name}")
